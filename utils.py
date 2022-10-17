from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl import Workbook


def format_sheet(sheet, dataframe):
    for r in dataframe_to_rows(dataframe, index=False, header=True):
        sheet.append(r)
    cells = [
        sheet["A1"],
        sheet["B1"],
        sheet["C1"],
        sheet["D1"],
        sheet["E1"],
        sheet["F1"],
        sheet["G1"],
        sheet["H1"],
        sheet["I1"],
        sheet["J1"],
        sheet["K1"],
        sheet["L1"],
    ]
    for cell in cells:
        cell.fill = PatternFill(
            start_color="E06666", end_color="E06666", fill_type="solid"
        )
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    cells = [
        sheet["M1"],
        sheet["N1"],
        sheet["O1"],
        sheet["P1"],
        sheet["Q1"],
        sheet["R1"],
        sheet["S1"],
        sheet["T1"],
        sheet["U1"],
        sheet["V1"],
        sheet["W1"],
        sheet["X1"],
        sheet["Y1"],
    ]
    for cell in cells:
        cell.fill = PatternFill(
            start_color="0070C0", end_color="0070C0", fill_type="solid"
        )
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet["Z1"].fill = PatternFill(
        start_color="60497A", end_color="60497A", fill_type="solid"
    )
    sheet["Z1"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    sheet["Z1"].alignment = Alignment(horizontal="center", vertical="center")

    # 60497A

    sheet.column_dimensions["A"].width = 16
    sheet.column_dimensions["B"].width = 7
    sheet.column_dimensions["C"].width = 25
    sheet.column_dimensions["D"].width = 12
    sheet.column_dimensions["E"].width = 43
    sheet.column_dimensions["F"].width = 19
    sheet.column_dimensions["G"].width = 36
    sheet.column_dimensions["H"].width = 15
    sheet.column_dimensions["H"].width = 24
    sheet.column_dimensions["I"].width = 25
    sheet.column_dimensions["J"].width = 25
    sheet.column_dimensions["K"].width = 28
    sheet.column_dimensions["L"].width = 25
    sheet.column_dimensions["M"].width = 21
    sheet.column_dimensions["N"].width = 13
    sheet.column_dimensions["O"].width = 26
    sheet.column_dimensions["P"].width = 11
    sheet.column_dimensions["Q"].width = 12
    sheet.column_dimensions["R"].width = 42
    sheet.column_dimensions["S"].width = 8
    sheet.column_dimensions["T"].width = 52
    sheet.column_dimensions["U"].width = 30
    sheet.column_dimensions["V"].width = 30
    sheet.column_dimensions["W"].width = 15
    sheet.column_dimensions["X"].width = 50
    sheet.column_dimensions["Y"].width = 10
    sheet.column_dimensions["Z"].width = 20

    for _ in range(2):
        sheet.insert_rows(1)


def get_status(x):
    if x["STATUS"] in ["Facturado", "Cerrado"] and x["mobile_status"] == "delivered":
        return "PERFECT"
    elif x["STATUS"] in ["Cerrado", "Cancelado"] and x["mobile_status"] == "canceled":
        return "OK Canceled"
    elif x["STATUS"] in ["Ejecución de la orden pendiente"] and x["mobile_status"] in [
        "order_confirmed",
        "order_preparation",
        "on_way_to_second_deliver",
        "on_route",
        "first_attempt_failure",
    ]:
        return "OK On-going"
    elif x["STATUS"] in ["Aprobación pendiente"] and x["mobile_status"] in [
        "pending_payment"
    ]:
        return "OK payment pending"
    elif x["STATUS"] in ["Ejecución de la orden pendiente"] and x["mobile_status"] in [
        "pending_payment"
    ]:
        return "OK payment pending"
    else:
        return "UNMATCHED"
